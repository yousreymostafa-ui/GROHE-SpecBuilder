from __future__ import annotations

import csv
import json
import os
import re
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

APP_ROOT = Path(__file__).resolve().parent
CONFIG_PATH = APP_ROOT / "grohe_selector_config.json"
STATE_PATH = APP_ROOT / "image_recovery_state.json"
DEFAULT_IMAGE_FOLDER = r"G:\My Drive\Images"
OFFICIAL_HOST_SUFFIXES = ("grohe.com", "grohe-mena.com", "cloud.grohe.com", "grohe.us", "lixil.cdn.celum.cloud")
IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp")
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/151 Safari/537.36"
STATE_LOCK = threading.Lock()


def normalize_sku(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", str(value or "")).upper()


def load_json(path: Path, default: Any) -> Any:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default


def save_json(path: Path, data: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def load_config() -> dict[str, Any]:
    cfg = load_json(CONFIG_PATH, {})
    if not isinstance(cfg, dict):
        cfg = {}
    env_folder = (os.environ.get("GROHE_IMAGES_DIR") or os.environ.get("GROHE_IMAGE_FOLDER") or "").strip()
    folder = env_folder or cfg.get("image_root") or cfg.get("image_folder") or cfg.get("images_folder") or cfg.get("imageFolder") or DEFAULT_IMAGE_FOLDER
    cfg["image_root"] = str(folder)
    cfg["image_folder"] = str(folder)
    return cfg


def save_config(patch: dict[str, Any]) -> None:
    existing = load_json(CONFIG_PATH, {})
    if not isinstance(existing, dict):
        existing = {}
    existing.update(patch)
    # Keep old/new selector keys synchronized.
    folder = existing.get("image_root") or existing.get("image_folder")
    if folder:
        existing["image_root"] = str(folder)
        existing["image_folder"] = str(folder)
    save_json(CONFIG_PATH, existing)


def image_folder() -> Path:
    return Path(load_config()["image_root"])


def set_image_folder(folder: str | Path) -> Path:
    p = Path(str(folder))
    save_config({"image_root": str(p), "image_folder": str(p)})
    return p


def load_state() -> dict[str, Any]:
    data = load_json(STATE_PATH, {})
    return data if isinstance(data, dict) else {}


def update_state(sku: str, patch: dict[str, Any]) -> dict[str, Any]:
    sku = normalize_sku(sku)
    with STATE_LOCK:
        data = load_state()
        current = data.get(sku, {}) if isinstance(data.get(sku), dict) else {}
        current.update(patch)
        current["sku"] = sku
        current["last_checked"] = int(time.time())
        data[sku] = current
        save_json(STATE_PATH, data)
        return current


def find_local_image(sku: str, folder: Path | None = None) -> Path | None:
    sku = normalize_sku(sku)
    folder = folder or image_folder()
    if not folder.exists():
        return None
    for ext in IMAGE_EXTS:
        for candidate in (folder / f"{sku}{ext}", folder / f"{sku}{ext.upper()}"):
            if candidate.exists() and candidate.is_file():
                return candidate
    try:
        for p in folder.iterdir():
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS and normalize_sku(p.stem) == sku:
                return p
    except Exception:
        return None
    return None


def official_url(url: str) -> bool:
    try:
        host = (urllib.parse.urlparse(url).hostname or "").lower()
        return any(host == suffix or host.endswith("." + suffix) for suffix in OFFICIAL_HOST_SUFFIXES)
    except Exception:
        return False


def _request_headers(*, image: bool = False) -> dict[str, str]:
    return {
        "User-Agent": USER_AGENT,
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8" if image else "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-GB,en;q=0.9",
        "Referer": "https://www.grohe-mena.com/",
        "Connection": "close",
    }


def request_url(url: str, *, timeout: int = 12, head: bool = False, max_bytes: int | None = None) -> tuple[int, dict[str, str], bytes]:
    headers = _request_headers(image=head or any(url.lower().split("?", 1)[0].endswith(ext) for ext in IMAGE_EXTS))
    req = urllib.request.Request(url, headers=headers, method="HEAD" if head else "GET")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            if head:
                body = b""
            elif max_bytes is not None:
                body = resp.read(max_bytes)
            else:
                body = resp.read()
            return int(getattr(resp, "status", 200)), {k.lower(): v for k, v in resp.headers.items()}, body
    except urllib.error.HTTPError as e:
        return int(e.code), {k.lower(): v for k, v in e.headers.items()} if e.headers else {}, b""
    except Exception as exc:
        return 0, {"x-error": str(exc)}, b""


def looks_like_image(headers: dict[str, str], body: bytes = b"", url: str = "") -> bool:
    ctype = headers.get("content-type", "").lower()
    if "image/" in ctype:
        return True
    if body.startswith(b"\xff\xd8\xff") or body.startswith(b"\x89PNG\r\n\x1a\n"):
        return True
    if len(body) >= 12 and body[:4] == b"RIFF" and body[8:12] == b"WEBP":
        return True
    if body[:6] in (b"GIF87a", b"GIF89a"):
        return True
    return False


def probe_image(url: str) -> tuple[bool, str]:
    """Probe an official image without relying on HEAD support.

    Some GROHE/CDN endpoints reject HEAD and Range requests even though a normal GET
    works. v15.2 therefore falls back to a small normal GET and records the reason
    when a candidate cannot be used.
    """
    if not official_url(url):
        return False, "URL is not on an approved GROHE host"
    status, headers, _ = request_url(url, timeout=7, head=True)
    if 200 <= status < 400 and looks_like_image(headers, b"", url):
        return True, ""
    status2, headers2, body2 = request_url(url, timeout=12, head=False, max_bytes=8192)
    if 200 <= status2 < 400 and looks_like_image(headers2, body2, url):
        return True, ""
    err = headers2.get("x-error") or headers.get("x-error") or f"HTTP {status2 or status or 0}"
    return False, err

def cdn_candidates(sku: str) -> list[str]:
    sku = normalize_sku(sku)
    m = re.match(r"^(\d{2})", sku)
    if not m:
        return []
    prefix2 = m.group(1)
    bucket = f"{(int(prefix2) // 10) * 10:02d}"
    base = f"https://cdn.cloud.grohe.com/prod/{bucket}/{prefix2}/{sku}"
    # GROHE's standard hero image is _1_1.jpg. Probe the high-probability paths
    # first so Find All does not make dozens of web requests for every SKU.
    rel = [
        (1280, f"{sku}_1_1.jpg"), (960, f"{sku}_1_1.jpg"), (480, f"{sku}_1_1.jpg"),
        (1280, f"{sku}_1_1.png"), (1280, f"{sku}_1_1.webp"),
        (1280, f"{sku}_1_2.jpg"), (1280, f"{sku}_1_3.jpg"),
        (960, f"{sku}_1_2.jpg"), (480, f"{sku}_1_2.jpg"), (1280, f"{sku}_2_1.jpg"),
    ]
    return [f"{base}/{size}/{name}" for size,name in rel]

def discover_mena_product_pages(sku: str) -> list[str]:
    """Best-effort official GROHE MENA product-page discovery by exact SKU."""
    sku = normalize_sku(sku)
    if not sku:
        return []
    queries = [
        f"https://www.grohe-mena.com/en_cy/search-results-page/?tab=productSearch&search={urllib.parse.quote(sku)}",
        f"https://www.grohe-mena.com/en_cy/search-results-page/?tab=productSearch&q={urllib.parse.quote(sku)}",
    ]
    found=[]
    for search_url in queries:
        status, _headers, body = request_url(search_url, timeout=12, head=False)
        if not (200 <= status < 400) or not body:
            continue
        text=body.decode("utf-8", errors="ignore").replace("\\/", "/")
        for href in re.findall(r'href=["\\\']([^"\\\']+)["\\\']', text, flags=re.I):
            url=urllib.parse.urljoin(search_url, href.replace("&amp;", "&"))
            if not official_url(url):
                continue
            if sku.lower() in url.lower() and url not in found:
                found.append(url)
        if found:
            break
    return found

def manifest_sources() -> dict[str, dict[str, str]]:
    candidates = [
        APP_ROOT / "GROHE_Download_Report.csv",
        APP_ROOT / "GROHE_Download_Report.xlsx",
        APP_ROOT.parent / "GROHE_Download_Report.csv",
        APP_ROOT.parent / "GROHE_Download_Report.xlsx",
    ]
    out: dict[str, dict[str, str]] = {}
    for path in candidates:
        if not path.exists():
            continue
        try:
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as f:
                    for row in csv.DictReader(f):
                        sku = normalize_sku(row.get("SKU", ""))
                        if sku:
                            out[sku] = {
                                "image_url": (row.get("Image source") or row.get("Image Source") or "").strip(),
                                "product_page": (row.get("Product page") or row.get("Product Page") or "").strip(),
                            }
            else:
                try:
                    from openpyxl import load_workbook  # type: ignore
                except Exception:
                    continue
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb["Download Report"] if "Download Report" in wb.sheetnames else wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                headers = [str(x or "").strip() for x in next(rows)]
                idx = {name: i for i, name in enumerate(headers)}
                for values in rows:
                    i_sku = idx.get("SKU")
                    sku = normalize_sku(values[i_sku] if i_sku is not None and i_sku < len(values) else "")
                    if not sku:
                        continue
                    def val(name: str) -> str:
                        i = idx.get(name)
                        return str(values[i] or "").strip() if i is not None and i < len(values) else ""
                    out[sku] = {"image_url": val("Image source"), "product_page": val("Product page")}
            if out:
                break
        except Exception:
            continue
    return out


def extract_page_candidates(page_url: str, sku: str) -> list[str]:
    if not official_url(page_url):
        return []
    status, _headers, body = request_url(page_url, timeout=12, head=False)
    if not (200 <= status < 400) or not body:
        return []
    text = body.decode("utf-8", errors="ignore").replace("\\/", "/")
    urls = re.findall(r'https?://[^"\'<>\s]+', text)
    cleaned: list[str] = []
    for raw in urls:
        url = raw.replace("&amp;", "&").rstrip("),;]")
        if "cdn.cloud.grohe.com" not in url.lower():
            continue
        low = url.lower().split("?", 1)[0]
        if any(low.endswith(ext) for ext in IMAGE_EXTS):
            cleaned.append(url)
    return sorted(set(cleaned), key=lambda u: (sku not in normalize_sku(u), len(u)))


def find_candidate(sku: str) -> dict[str, Any]:
    sku = normalize_sku(sku)
    local = find_local_image(sku)
    if local:
        return update_state(sku, {"status": "local", "local_file": str(local), "confidence": 100, "source": "Local image folder", "last_error": ""})

    attempts: list[str] = []
    last_error = ""
    manifest = manifest_sources().get(sku, {})
    manifest_url = manifest.get("image_url", "")
    product_page = manifest.get("product_page", "")
    if manifest_url and official_url(manifest_url):
        ok, err = probe_image(manifest_url)
        attempts.append(manifest_url)
        if ok:
            return update_state(sku, {
                "status": "confirmed", "source_url": manifest_url, "product_page": product_page,
                "confidence": 100, "source": "GROHE download report", "attempts": attempts[-8:], "last_error": ""
            })
        last_error = err

    # Fast official-CDN route. This is also the path used in GROHE's existing
    # download reports (e.g. /prod/30/32/32109001/1280/32109001_1_1.jpg).
    for url in cdn_candidates(sku):
        ok, err = probe_image(url)
        attempts.append(url)
        if ok:
            return update_state(sku, {
                "status": "confirmed", "source_url": url,
                "product_page": product_page or f"https://www.grohe.com/en-GB/product/{sku}",
                "confidence": 100, "source": "GROHE official CDN", "attempts": attempts[-8:], "last_error": ""
            })
        last_error = err

    # GROHE MENA is usually easier to access from IMEA than the new global site.
    discovered = discover_mena_product_pages(sku)
    pages = [product_page, *discovered, f"https://www.grohe.com/en-GB/product/{sku}"]
    seen_pages: set[str] = set()
    for page in pages:
        if not page or page in seen_pages or not official_url(page):
            continue
        seen_pages.add(page)
        for url in extract_page_candidates(page, sku):
            ok, err = probe_image(url)
            attempts.append(url)
            if not ok:
                last_error = err
                continue
            exact = sku in normalize_sku(url)
            return update_state(sku, {
                "status": "confirmed" if exact else "review",
                "source_url": url, "product_page": page,
                "confidence": 100 if exact else 70,
                "source": "GROHE official product page", "attempts": attempts[-8:], "last_error": ""
            })

    reason = last_error or "No exact official image found"
    return update_state(sku, {
        "status": "missing", "confidence": 0, "source": "No exact official image found",
        "last_error": reason, "attempts": attempts[-8:]
    })

def download_candidate(sku: str, url: str | None = None) -> dict[str, Any]:
    sku = normalize_sku(sku)
    state = load_state().get(sku, {})
    url = (url or (state.get("source_url") if isinstance(state, dict) else "") or "").strip()
    if not url or not official_url(url):
        raise ValueError("No approved official GROHE image URL available. Run Find first.")
    status, headers, body = request_url(url, timeout=35, head=False)
    if not (200 <= status < 400) or not body or not looks_like_image(headers, body[:32], url):
        reason=headers.get("x-error") or f"HTTP {status or 0}"
        update_state(sku,{"last_error":reason})
        raise ValueError(f"The GROHE image could not be downloaded ({reason})")
    ctype = headers.get("content-type", "").lower()
    ext = ".jpg"
    if "png" in ctype or body.startswith(b"\x89PNG"):
        ext = ".png"
    elif "webp" in ctype or (len(body)>=12 and body[:4]==b"RIFF" and body[8:12]==b"WEBP"):
        ext = ".webp"
    elif "gif" in ctype or body[:6] in (b"GIF87a",b"GIF89a"):
        ext = ".gif"
    folder = image_folder()
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / f"{sku}{ext}"
    target.write_bytes(body)
    return update_state(sku, {
        "status": "local", "local_file": str(target), "source_url": url,
        "confidence": 100, "downloaded_at": int(time.time()), "last_error": "",
        "source": state.get("source", "GROHE official") if isinstance(state, dict) else "GROHE official"
    })

def scan_products(skus: list[str], local_index: dict[str, Path] | None = None) -> dict[str, Any]:
    state = load_state()
    items: dict[str, Any] = {}
    local_count = 0
    folder = image_folder()
    for raw in skus:
        sku = normalize_sku(raw)
        if not sku:
            continue
        p = local_index.get(sku) if local_index is not None else find_local_image(sku, folder)
        if p:
            local_count += 1
            items[sku] = {"status": "local", "local_file": str(p), "confidence": 100, "source": "Local image folder"}
        else:
            old = state.get(sku, {}) if isinstance(state.get(sku), dict) else {}
            items[sku] = old if old else {"status": "unchecked", "confidence": 0}
    return {
        "ok": True, "items": items, "total": len(items), "local": local_count,
        "missing": max(0, len(items) - local_count), "image_folder": str(folder),
        "folder_exists": folder.exists()
    }
